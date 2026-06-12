#!/usr/bin/env python3
"""Turn a set of vless:// lines into shareable PNG / HTML / PDF / XLSX exports.

This is the *one* exporter shared by all three VPN-line skills
(reality-wireguard-relay, xray-relay-deploy, xray-xhttp-relay-deploy). Every one
of those skills ends up with a plain-text blob full of ``vless://`` links — the
relay skill's ``all-links.txt``, or the ``/root/xray_nodes_info.txt`` the
one-click scripts write. The links are the single source of truth: address,
port, transport and REALITY params all live inside the URL, so this script reads
*only* the ``vless://`` lines and ignores whatever decoration surrounds them.
That keeps one parser working across all three input shapes.

    pip install --user segno reportlab openpyxl pillow   # one-time
    # links already on disk:
    python3 export_links.py --input all-links.txt --out ./exports
    # or pipe them in (e.g. straight off a VPS):
    ssh myvps 'cat /root/xray_nodes_info.txt' | python3 export_links.py --out ./exports

Formats (default: all four). Each degrades on its own — a missing library skips
just that format with a pip hint, the others still run:

  png   one QR per line, <name>.png            (segno)
  html  self-contained gallery, QR inlined      (no deps)
  pdf   printable sheet, one line per row        (reportlab)
  xlsx  spreadsheet table + embedded QR images   (openpyxl + pillow)

The QR matters because IPv6 brackets, ``#`` and ``&`` in a vless link get
mangled when copy-pasted through chat apps — scanning the PNG sidesteps that.
"""
import argparse
import io
import os
import re
import sys
from urllib.parse import urlsplit, parse_qs, unquote

VLESS_RE = re.compile(r"vless://\S+")


def parse_links(text):
    """Pull every vless:// link out of arbitrary text into structured records.

    Order is preserved and exact duplicates are dropped (the xray info files
    sometimes echo a link in more than one section)."""
    records, seen = [], set()
    for m in VLESS_RE.finditer(text):
        link = m.group(0).rstrip(".,)】」>\"'")  # trim trailing punctuation
        if link in seen:
            continue
        seen.add(link)
        u = urlsplit(link)
        q = parse_qs(u.query)
        name = unquote(u.fragment) or f"line-{len(records) + 1}"
        records.append({
            "name": name,
            "link": link,
            "host": u.hostname or "",
            "port": str(u.port or ""),
            "transport": (q.get("type") or [""])[0],
            "security": (q.get("security") or [""])[0],
            "sni": (q.get("sni") or [""])[0],
        })
    return records


def safe_filename(name):
    return re.sub(r"[^\w.\-]+", "_", name).strip("_") or "line"


def qr_png_bytes(link, scale=8, border=3):
    import segno
    buf = io.BytesIO()
    segno.make(link, error="m").save(buf, kind="png", scale=scale, border=border)
    return buf.getvalue()


# ---- exporters -------------------------------------------------------------

def export_png(records, out, qr_cache):
    qr_dir = os.path.join(out, "qr")
    os.makedirs(qr_dir, exist_ok=True)
    for rec in records:
        path = os.path.join(qr_dir, safe_filename(rec["name"]) + ".png")
        with open(path, "wb") as fh:
            fh.write(qr_cache[rec["link"]])
    return f"{len(records)} QR PNG(s) in {qr_dir}/"


def export_html(records, out, qr_cache):
    import base64
    cards = []
    for rec in records:
        b64 = base64.b64encode(qr_cache[rec["link"]]).decode()
        meta = " · ".join(filter(None, [
            f"port {rec['port']}" if rec["port"] else "",
            rec["host"],
            rec["transport"],
            rec["security"],
        ]))
        cards.append(f"""    <div class="card">
      <img alt="QR for {rec['name']}" src="data:image/png;base64,{b64}">
      <div class="info">
        <h2>{_esc(rec['name'])}</h2>
        <p class="meta">{_esc(meta)}</p>
        <code>{_esc(rec['link'])}</code>
      </div>
    </div>""")
    html = f"""<!doctype html>
<html lang="zh"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>VPN 线路 · {len(records)} 条</title>
<style>
  body {{ font-family: -apple-system, system-ui, sans-serif; margin: 2rem auto; max-width: 820px; color: #1a1a1a; }}
  h1 {{ font-size: 1.3rem; }}
  .card {{ display: flex; gap: 1rem; align-items: center; border: 1px solid #e3e3e3;
          border-radius: 12px; padding: 1rem; margin: 0.8rem 0; }}
  .card img {{ width: 150px; height: 150px; flex: none; }}
  .info {{ min-width: 0; }}
  .info h2 {{ font-size: 1.05rem; margin: 0 0 0.2rem; }}
  .meta {{ color: #666; font-size: 0.85rem; margin: 0 0 0.5rem; }}
  code {{ display: block; word-break: break-all; font-size: 0.8rem; background: #f6f6f6;
          padding: 0.5rem; border-radius: 6px; user-select: all; }}
</style></head><body>
<h1>VPN 线路 · 共 {len(records)} 条</h1>
{chr(10).join(cards)}
</body></html>"""
    path = os.path.join(out, "lines.html")
    with open(path, "w", encoding="utf-8") as fh:
        fh.write(html)
    return f"HTML gallery → {path}"


def _esc(s):
    return (s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


def export_pdf(records, out, qr_cache):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                    TableStyle, Image)
    from reportlab.lib import colors

    styles = getSampleStyleSheet()
    name_style = ParagraphStyle("name", parent=styles["Heading3"], spaceAfter=2)
    meta_style = ParagraphStyle("meta", parent=styles["Normal"], textColor=colors.grey,
                                fontSize=8, spaceAfter=4)
    # wordWrap=CJK breaks anywhere, which a long space-less vless URL needs.
    link_style = ParagraphStyle("link", parent=styles["Code"], fontSize=7,
                                wordWrap="CJK", leading=9)

    path = os.path.join(out, "lines.pdf")
    doc = SimpleDocTemplate(path, pagesize=A4, title="VPN lines",
                            leftMargin=15 * mm, rightMargin=15 * mm,
                            topMargin=15 * mm, bottomMargin=15 * mm)
    flow = [Paragraph(f"VPN 线路 · 共 {len(records)} 条", styles["Title"]),
            Spacer(1, 6 * mm)]
    for rec in records:
        meta = " · ".join(filter(None, [
            f"port {rec['port']}" if rec["port"] else "", rec["host"],
            rec["transport"], rec["security"]]))
        qr = Image(io.BytesIO(qr_cache[rec["link"]]), width=32 * mm, height=32 * mm)
        text = [Paragraph(_esc(rec["name"]), name_style),
                Paragraph(_esc(meta), meta_style),
                Paragraph(_esc(rec["link"]), link_style)]
        row = Table([[qr, text]], colWidths=[35 * mm, None])
        row.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LINEBELOW", (0, 0), (-1, -1), 0.4, colors.HexColor("#e3e3e3")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
        ]))
        flow.append(row)
    doc.build(flow)
    return f"PDF → {path}"


def export_xlsx(records, out, qr_cache):
    import openpyxl
    from openpyxl.styles import Font, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VPN 线路"
    headers = ["名称", "QR", "地址", "端口", "传输", "安全", "vless 链接"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
    widths = [22, 24, 26, 8, 10, 10, 70]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    try:
        from openpyxl.drawing.image import Image as XLImage
        have_img = True
    except Exception:
        have_img = False  # pragma: no cover

    for r, rec in enumerate(records, start=2):
        ws.cell(r, 1, rec["name"])
        ws.cell(r, 3, rec["host"])
        ws.cell(r, 4, rec["port"])
        ws.cell(r, 5, rec["transport"])
        ws.cell(r, 6, rec["security"])
        cell = ws.cell(r, 7, rec["link"])
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if have_img:
            try:
                img = XLImage(io.BytesIO(qr_cache[rec["link"]]))
                img.width = img.height = 120
                ws.row_dimensions[r].height = 95
                ws.add_image(img, f"B{r}")
            except Exception:
                have_img = False  # Pillow missing → fall back to text-only
    path = os.path.join(out, "lines.xlsx")
    wb.save(path)
    note = "" if have_img else "  (QR images skipped — pip install --user pillow)"
    return f"XLSX → {path}{note}"


EXPORTERS = {"png": export_png, "html": export_html, "pdf": export_pdf, "xlsx": export_xlsx}
PIP_HINT = {"png": "segno", "html": "(none)", "pdf": "reportlab", "xlsx": "openpyxl pillow"}


def main():
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--input", help="file with vless:// links (default: stdin)")
    ap.add_argument("--out", default="./exports", help="output directory")
    ap.add_argument("--format", default="all",
                    help="comma list of png,html,pdf,xlsx (default: all)")
    args = ap.parse_args()

    text = open(args.input, encoding="utf-8").read() if args.input else sys.stdin.read()
    records = parse_links(text)
    if not records:
        sys.exit("no vless:// links found in input")

    fmts = list(EXPORTERS) if args.format == "all" else [
        f.strip() for f in args.format.split(",") if f.strip()]
    bad = [f for f in fmts if f not in EXPORTERS]
    if bad:
        sys.exit(f"unknown format(s): {', '.join(bad)} (pick from {', '.join(EXPORTERS)})")

    os.makedirs(args.out, exist_ok=True)

    # QR bytes are shared by every format; build once.
    qr_cache = {}
    if any(f in fmts for f in ("png", "html", "pdf", "xlsx")):
        try:
            for rec in records:
                qr_cache[rec["link"]] = qr_png_bytes(rec["link"])
        except ImportError:
            sys.exit("segno not installed (pip install --user segno) — needed for QR codes")

    print(f"parsed {len(records)} line(s); writing to {args.out}/")
    for f in fmts:
        try:
            print("  " + EXPORTERS[f](records, args.out, qr_cache))
        except ImportError:
            print(f"  [{f}] skipped — pip install --user {PIP_HINT[f]}", file=sys.stderr)


if __name__ == "__main__":
    main()
